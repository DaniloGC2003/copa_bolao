import pandas as pd
import streamlit as st
import zipfile
import tempfile
from pathlib import Path
from openpyxl import load_workbook
import math

NUMERO_REGRAS = 6
NUMERO_PARTIDAS_GRUPOS = 72
CONDICAO_ACERTOU_PLACAR = "Acertou placar"
CONDICAO_ACERTOU_VENCEDOR = "Acertou vencedor"
CONDICAO_ACERTOU_EMPATE = "Acertou empate"
MULTIPLICADOR_TERCEIRO_FINAL = "Multiplicador terceiro lugar e final"
MULTIPLICADOR_SEMIFINAIS = "Multiplicador semifinais"
MULTIPLICADOR_QUARTAS = "Multiplicador quartas"
EMPATE = "EMPATE"
POSICAO_COLUNA_PONTOS_FASE_GRUPOS = 9
POSICAO_COLUNA_PONTOS_ELIMINATORIAS = 11
regras = []
pontuacoes_participantes = []

# Ler arquivo de regras
df = pd.read_excel(
    "regras_pontuacao.xlsx",
    engine="openpyxl"
)
# Extrair regras
for it in range(NUMERO_REGRAS):
    regras.append({"regra":  df.iloc[it+2,0],
                   "pontos": df.iloc[it+2,1]})
pontos_acertou_placar = None
pontos_acertou_vencedor = None
pontos_acertou_empate = None
multiplicador_terceiro_final = None
multiplicador_semifinais = None
multiplicador_quartas = None

for d in regras:
    if d["regra"] == CONDICAO_ACERTOU_PLACAR:
        pontos_acertou_placar = d["pontos"]
    if d["regra"] == CONDICAO_ACERTOU_VENCEDOR:
        pontos_acertou_vencedor = d["pontos"]
    if d["regra"] == CONDICAO_ACERTOU_EMPATE:
        pontos_acertou_empate = d["pontos"]
    if d["regra"] == MULTIPLICADOR_TERCEIRO_FINAL:
        multiplicador_terceiro_final = d["pontos"]
    if d["regra"] == MULTIPLICADOR_SEMIFINAIS:
        multiplicador_semifinais = d["pontos"]
    if d["regra"] == MULTIPLICADOR_QUARTAS:
        multiplicador_quartas = d["pontos"]

def exec_round(gabarito, diretorio_participantes, numero_partidas, multiplicador, coluna_pontos):
    # Ler arquivo de resultado oficial
    file = pd.read_excel(
        gabarito,
        engine="openpyxl"
    )
    # Extrair resultados oficiais
    partidas = []
    for i in range(numero_partidas):
        partida_atual = {"equipe_1": file.iloc[i + 3, 3],
                         "equipe_2": file.iloc[i + 3, 4],
                         "score_1": file.iloc[i + 3, 5],
                         "score_2": file.iloc[i + 3, 6],
                         "vencedor": 0
                         }
        if math.isnan(partida_atual["score_1"]) or math.isnan(partida_atual["score_2"]):
            partida_atual["vencedor"] = "?"
        elif partida_atual["score_1"] > partida_atual["score_2"]:
            partida_atual["vencedor"] = partida_atual["equipe_1"]
        elif partida_atual["score_2"] > partida_atual["score_1"]:
            partida_atual["vencedor"] = partida_atual["equipe_2"]
        elif partida_atual["score_1"] == partida_atual["score_2"]:
            if numero_partidas == NUMERO_PARTIDAS_GRUPOS:
                partida_atual["vencedor"] = "EMPATE"
            else:
                if math.isnan(file.iloc[i + 3, 8]) and (file.iloc[i + 3, 9] == "x" or file.iloc[i + 3, 9] == "X"):
                    partida_atual["vencedor"] = partida_atual["equipe_2"]
                elif math.isnan(file.iloc[i + 3, 9]) and (file.iloc[i + 3, 8] == "x" or file.iloc[i + 3, 8] == "X"):
                    partida_atual["vencedor"] = partida_atual["equipe_1"]
                else:
                    partida_atual["vencedor"] = "?"
        partidas.append(partida_atual)
    print("\tpartidas na rodada: ")
    for partida in partidas:
       print(f"\t{partida}")

    # Contar pontos de cada participante
    pasta = Path(diretorio_participantes)

    for arquivo in pasta.iterdir():
        if arquivo.suffix == ".xlsx":
            print(f"\t{arquivo}")
            pontuacao_total = 0
            file = pd.read_excel(
                arquivo,
                engine="openpyxl"
            )
            workbook = load_workbook(arquivo)
            print("SHEETS: ")
            print(workbook.sheetnames)
            sheet = workbook["Palpites"]
            nome_participante = file.iloc[0, 2]
            print(f"\tContando pontos de {nome_participante}")
            for i in range(numero_partidas):
                previsao_participante = {
                    "equipe_1": file.iloc[i + 3, 3],
                    "equipe_2": file.iloc[i + 3, 4],
                    "score_1": file.iloc[i + 3, 5],
                    "score_2": file.iloc[i + 3, 6],
                    "vencedor": 0
                }
                if math.isnan(previsao_participante["score_1"]) or math.isnan(previsao_participante["score_2"]):
                    previsao_participante["vencedor"] = "?"
                elif previsao_participante["score_1"] > previsao_participante["score_2"]:
                    previsao_participante["vencedor"] = previsao_participante["equipe_1"]
                elif previsao_participante["score_2"] > previsao_participante["score_1"]:
                    previsao_participante["vencedor"] = previsao_participante["equipe_2"]
                elif previsao_participante["score_1"] == previsao_participante["score_2"]:
                    if numero_partidas == NUMERO_PARTIDAS_GRUPOS:
                        previsao_participante["vencedor"] = "EMPATE"
                    else:
                        if pd.isna(file.iloc[i + 3, 8]) and (file.iloc[i + 3, 9] == "x" or file.iloc[i + 3, 9] == "X"):
                            previsao_participante["vencedor"] = previsao_participante["equipe_2"]
                        elif pd.isna(file.iloc[i + 3, 9]) and (file.iloc[i + 3, 8] == "x" or file.iloc[i + 3, 8] == "X"):
                            previsao_participante["vencedor"] = previsao_participante["equipe_1"]
                        else:
                            previsao_participante["vencedor"] = "?"

                print("previsao: ")
                print(previsao_participante)
                # Valores para vencedor:
                # "?": resultados nao preenchidos corretamente
                # "EMPATE": empate
                # TIME: nome do time vencedor
                if partidas[i]["vencedor"] != "?":
                    if previsao_participante["vencedor"] != "?":
                        pontos_na_partida = 0
                        acertou = False
                        print(f"\tAnalisando {partidas[i]["equipe_1"]} x {partidas[i]["equipe_2"]}")

                        # Caso participante tenha acertado placar
                        if previsao_participante["score_1"] == partidas[i]["score_1"] and previsao_participante[
                            "score_2"] == partidas[i]["score_2"]:
                            pontuacao_total += pontos_acertou_placar * multiplicador
                            acertou = True
                            print("\t\tacertou placar")
                            pontos_na_partida += 3
                        # Caso participante tenha acertado vencedor
                        if previsao_participante["vencedor"] == partidas[i]["vencedor"] and partidas[i][
                            "vencedor"] != EMPATE:
                            pontuacao_total += pontos_acertou_vencedor * multiplicador
                            acertou = True
                            print("\t\tacertou vencedor")
                            pontos_na_partida += 3
                        # Caso participante tenha acertado empate
                        if previsao_participante["vencedor"] == partidas[i]["vencedor"] and partidas[i][
                            "vencedor"] == EMPATE:
                            pontuacao_total += pontos_acertou_empate * multiplicador
                            acertou = True
                            print("\t\tacertou empate")
                            pontos_na_partida += 3
                        if not acertou:
                            print("\t\tnao acertou")
                        print("escrevendo pontos na planilha")
                        sheet.cell(row=i+5, column=coluna_pontos).value = pontos_na_partida
                    else:
                        print(
                            f"\tParticipante nao preencheu {partidas[i]["equipe_1"]} x {partidas[i]["equipe_2"]} corretamente")
                        sheet.cell(row=i + 5, column=coluna_pontos).value = 0
                else:
                    print(
                        f"\tResultado oficial de {partidas[i]["equipe_1"]} x {partidas[i]["equipe_2"]} nao preenchido corretamente")
                    sheet.cell(row=i + 5, column=coluna_pontos).value = 0
            workbook.save(arquivo)
            # Adiciona pontuacao a pontuacoes_participantes. Cria nova entrada caso nao exista
            achou_participante = False
            for participante in pontuacoes_participantes:
                if participante["nome_participante"] == nome_participante:
                    #print("Participante ja existe")
                    participante["pontuacao_total"] += pontuacao_total
                    achou_participante = True
            if not achou_participante:
                #print("Criando novo participante")
                pontuacoes_participantes.append({"nome_participante": nome_participante,
                                                "pontuacao_total": pontuacao_total})
        print()
    print("Pontuacoes apos essa rodada: ")
    print(pontuacoes_participantes)
    print()


print("EXEC FASE DE GRUPOS")
exec_round("resultado_oficial_fase_grupos.xlsx", "planilhas_participantes/fase_de_grupos",
           NUMERO_PARTIDAS_GRUPOS, 1, POSICAO_COLUNA_PONTOS_FASE_GRUPOS)
print("EXEC FASE DE 32")
exec_round("resultado_oficial_rodada_32.xlsx", "planilhas_participantes/fase_32",
           16, 1, POSICAO_COLUNA_PONTOS_ELIMINATORIAS)
print("EXEC OITAVAS")
exec_round("resultado_oficial_oitavas.xlsx", "planilhas_participantes/oitavas",
           8, 1, POSICAO_COLUNA_PONTOS_ELIMINATORIAS)
print("EXEC QUARTAS")
exec_round("resultado_oficial_quartas.xlsx", "planilhas_participantes/quartas",
           4, 1, POSICAO_COLUNA_PONTOS_ELIMINATORIAS)
print("EXEC SEMIFINAIS")
exec_round("resultado_oficial_semifinais.xlsx", "planilhas_participantes/semifinais",
           2, 1, POSICAO_COLUNA_PONTOS_ELIMINATORIAS)
print("EXEC TERCEIRO LUGAR E FINAL")
exec_round("resultado_oficial_terceiro_lugar_final.xlsx", "planilhas_participantes/terceiro_lugar_final",
           2, multiplicador_terceiro_final, POSICAO_COLUNA_PONTOS_ELIMINATORIAS)




print("pontuacoes participantes: ")
print(pontuacoes_participantes)
pontuacoes_participantes_ordenado = sorted(pontuacoes_participantes, key=lambda x: x["pontuacao_total"], reverse=True)
print("pontuacoes participantes ordenado: ")
print(pontuacoes_participantes_ordenado)

# Criar planilha com pontuacao final
df = pd.DataFrame(pontuacoes_participantes_ordenado)
df.to_excel(
    "pontuacao_final.xlsx",
    engine="openpyxl",
    index=False
)

print()
input("Press Enter to exit...")
